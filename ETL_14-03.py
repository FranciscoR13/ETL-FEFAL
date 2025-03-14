from rapidfuzz import fuzz
from unidecode import unidecode
import pandas as pd
import dataframe_image as dfi
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from spellchecker import SpellChecker  
import psutil
import json
import re
import os

#================================================ FUNÇÕES ===============================================#

# Função para fechar o ficheiro de destino caso esteja aberto
def close_excel():
    for process in psutil.process_iter(attrs=["pid", "name"]):
        if "EXCEL.EXE" in process.info["name"]:                     # Verifica se o Excel está aberto
            os.system(f"taskkill /PID {process.info['pid']} /F")    # Força o encerramento
            return

# Função para corrigir erros ortográficos
def correct_text(texto):
    palavras = texto.split()
    palavras_corrigidas = [spell.correction(palavra) if spell.correction(palavra) else palavra for palavra in palavras]
    return " ".join(palavras_corrigidas).upper()                                                                         

# Função para normalizar os dados
def normalize_text(texto):
    if not isinstance(texto, str) or not texto.strip():  
        return ""  
    
    texto = texto.strip()                # Remover espaços no início e no fim
    texto = unidecode(texto)             # Remover acentos
    texto = re.sub(r"\s+", " ", texto)   # Substituir múltiplos espaços por um único espaço
    texto = texto.upper()                # Converter para maiúsculas
    
    return texto
    
# Função para normalizar texto e remover prefixos
def clean_text(text):
    if not isinstance(text, str):
        return text
    text = unidecode(text).strip().upper()
    for prefix in remover_prefixos:
        text = re.sub(prefix, "", text)
    return text

# Função para encontrar a melhor correspondência para um nome de coluna
def find_best_match(nome_coluna, lista_colunas, score_minimo=80):
    nome_coluna_norm = normalize_text(nome_coluna)  

    correspondencias = [
        (col, fuzz.partial_ratio(nome_coluna_norm, normalize_text(col))) 
        for col in lista_colunas
    ]
    
    correspondencias.sort(key=lambda x: x[1], reverse=True)  
    
    melhor_correspondencia, melhor_score = correspondencias[0]


    return melhor_correspondencia if melhor_score >= score_minimo else None

# Função para encontrar a melhor correspondência de coluna
def get_best_column(df, column_name, aliases, score_minimo=80):
    # Primeiro, verifica se a coluna principal existe diretamente
    if column_name in df.columns:
        return column_name

    # Se a coluna não existir diretamente, verificar os aliases
    for alias in aliases.get(column_name, []):
        # Tentar encontrar a melhor correspondência de fuzzy matching para cada alias
        best_match_column = find_best_match(alias, df.columns, score_minimo)

        if best_match_column:
            return list(aliases.keys())[0]

    return None

# Função de fuzzy matching genérica para concelhos e freguesias
def fuzzy_match_local(valor, local_list, limite_score=90):
    # Verifica se o valor tem uma correspondência fuzzy suficiente com qualquer localidade da lista
    for local in local_list:
        if fuzz.partial_ratio(valor, local) > limite_score:
            return local
    return None


#========================================================================================================#

#================================================ INICIO ================================================#

# Carregar JSON
with open('config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)

# Aceder às variáveis do JSON
ANO = config["ano"]
FILE_PATH_IN = config["file_paths"]["input"].format(ano=ANO)
file_path_out = config["file_paths"]["output"].format(ano=ANO)
file_path_removed = config["file_paths"]["removed"].format(ano=ANO)
file_path_concelhos = config["file_paths"]["concelhos"]
file_path_freguesias = config["file_paths"]["freguesias"]

cols_targets = config["columns"]["targets"]
aliases = config["columns"].get("aliases", {})
entidade_type = config["columns"]["entity_type"]
responsible = config["columns"]["responsible"]
col_nformandos = config["columns"]["num_formandos"]

# Ler o ficheiro Excel
df = pd.read_excel(FILE_PATH_IN)

# Normalizar os nomes das colunas
# Aplicar a normalização a todas as colunas
df.columns = [normalize_text(col) for col in df.columns]

# Identificar a melhor coluna para verificar duplicados
coluna_verificar = get_best_column(df, config["columns"]["check_duplicates"], aliases)
if not coluna_verificar:
    raise ValueError("Erro: Nenhuma correspondência encontrada para 'check_duplicates'.")

valores_invalidos = config["invalid_values"]
ws_title = config["ws_title"].format(ano=ANO)
limite_fuzzy = config["fuzzy_limit"]

n_entidades_municipios = config["keywords"]["n_entidades_municipios"]
n_entidades_freguesias = config["keywords"]["n_entidades_freguesias"]
municipio_keywords = config["keywords"]["municipio"]
freguesia_keywords = config["keywords"]["freguesia"]
tipos_entidade_validos = config["keywords"]["entity_types"]

palavras_chave_formacao = config["keywords"]["training"]
palavras_chave_comentario = config["keywords"]["comment"]
palavras_chave_tempo_grupo = config["keywords"]["group_time"]
palavras_chave_areas_tematicas = config["keywords"]["thematic_areas"]
palavras_chave_continua = config["keywords"]["continuous_training"]
palavras_chave_preferencia = config["keywords"]["preference"]
palavras_chave_regime = config["keywords"]["regime"]

descricao_comentario = config["descriptions"]["comment"]
descricao_tempo_grupo = config["descriptions"]["group_time"]
descricao_interesse = config["descriptions"]["interest"]
descricao_continua = config["descriptions"]["continuous_training"]
descricao_preferencia = config["descriptions"]["preference"]
descricao_formacao_curso = config["descriptions"]["training_course"]
descricao_regime = config["descriptions"]["regime"]

data_sub_key = config["data_keys"]["submission_date"]
data_inicio_key = config["data_keys"]["start_date"]
data_fim_key = config["data_keys"]["end_date"]
col_sub = config["data_keys"]["submitted"]
col_temp = config["data_keys"]["completion_time"]

VALOR_VAZIO = config["default_values"]["empty"]
VALOR_NAO = config["default_values"]["no"]

trainings = config["trainings"]
interests = config["interests"]

remover_prefixos = config["prefixs"]

# Criar um dicionário de mapeamento {nome_original: nome_modificado}
mapeamento_colunas = {}
cols_comentarios_fc = []
cols_interesses_fc = []
cols_continua_fc = []
cols_sumformados = []

# Lista de Concelhos
with open(file_path_concelhos, 'r', encoding='utf-8') as file:
    concelhos = file.read()
concelhos = unidecode(concelhos)
concelhos_list = concelhos.split('\n')

# Lista de Freguesias
with open(file_path_freguesias, 'r', encoding='utf-8') as file:
    freguesias = file.read()
freguesias = unidecode(freguesias)
freguesias_list = freguesias.split('\n')

spell = SpellChecker(language="pt")  

# Criar um DataFrame para armazenar as linhas removidas
df_removidos = pd.DataFrame()

#========================================================================================================#

#=========================================== PREPARAR DATASET ===========================================#

# Fechar o Excel de destino caso esteja aberto
close_excel()

# Ler o ficheiro Excel
df = pd.read_excel(FILE_PATH_IN)

# Normalizar os nomes das colunas do DataFrame
df.columns = [normalize_text(col) for col in df.columns]

# 1. Colunas com apenas valores inválidos
colunas_invalidas = df.columns[df.apply(lambda col: col.isin(valores_invalidos).all(), axis=0)]

# 2. Colunas com apenas valores nulos
colunas_nulas = df.columns[df.isna().all()]

# 3. Combinar ambas corretamente
colunas_a_remover = colunas_invalidas.union(colunas_nulas)

df = df.drop(columns=colunas_a_remover)

# Aplicar normalize_text apenas em colunas de texto
for col in df.select_dtypes(include=['object', 'string']).columns:
    df[col] = df[col].apply(normalize_text)

# Dicionário para armazenar a melhor correspondência de cada coluna-alvo
best_matches = {}

municipio = False

# Encontrar a melhor correspondência para cada coluna-alvo
for target in cols_targets:
    best_match = find_best_match(target, df.columns, 90)

    if best_match:
        best_matches[best_match] = target
        continue  

    if target in aliases:
        for alias in aliases[target]:
            alias_match = find_best_match(alias, df.columns)
            if alias_match and alias_match not in best_matches:
                best_matches[alias_match] = target
                municipio = True

# Renomear as colunas do DataFrame
df = df.rename(columns=best_matches)

# Criar a coluna com valor "MUNICIPIO" se necessário
if municipio and "ENTIDADE DO SUBSETOR DA ADMINISTRAÇÃO LOCAL" not in df.columns:
    df["ENTIDADE DO SUBSETOR DA ADMINISTRAÇÃO LOCAL"] = "MUNICIPIO"

# Verificar se a coluna existe
if coluna_verificar in df.columns:
    # Criar máscara booleana para valores inválidos e nulos
    mask_valores_invalidos = df[coluna_verificar].isin(valores_invalidos)
    mask_nulos = df[coluna_verificar].isna()
    mask_remocao = mask_valores_invalidos | mask_nulos

    if mask_remocao.any():
        # Guardar as linhas removidas com motivo
        removidos = df[mask_remocao].copy()
        removidos["MOTIVO REMOCAO"] = "VALOR DE ENTIDADE NULO"
        # Garantir que a coluna do motivo esteja no início
        cols = ["MOTIVO REMOCAO"] + [col for col in removidos.columns if col != "MOTIVO REMOCAO"]
        removidos = removidos[cols]

        # Concatenar com as outras já removidas
        df_removidos = pd.concat([df_removidos, removidos], ignore_index=True)

        # Remover as linhas do DataFrame original
        df = df[~mask_remocao]
else:
    print(f"Aviso: A coluna '{coluna_verificar}' não existe no DataFrame. Nenhuma remoção foi feita.")


#========================================================================================================#

#=========================================== VALIDA ENTIDADES ===========================================#

entidades_invalidas = []

def validar_local(valor_verificado, local_list, limite_fuzzy):

    if not valor_verificado or len(normalize_text(valor_verificado)) <= 4:
        return None  

    # Tentativa de correspondência exata na lista de concelhos
    valor_verificado_norm = find_best_match(clean_text(valor_verificado), local_list, limite_fuzzy)
    if local_list and valor_verificado_norm:
        return clean_text(valor_verificado)

    # Correspondência difusa (fuzzy match)
    local_correspondente = fuzzy_match_local(clean_text(valor_verificado), local_list, limite_fuzzy) if local_list else None
    if local_correspondente:
        return clean_text(local_correspondente)

    return None  # Se nenhuma correspondência foi encontrada

def validar_generico_municipio_freguesia(valor_verificado):

    if not valor_verificado or len(normalize_text(valor_verificado)) <= 4:
        return None  

    valor_verificado_limpo = clean_text(valor_verificado)

    # Verifica se contém alguma palavra-chave de município ou freguesia
    if any(keyword in valor_verificado_limpo for keyword in municipio_keywords + freguesia_keywords):
        return valor_verificado_limpo

    return None  

def processar_entidade(row, coluna_verificar, entidade_type, concelhos_list, freguesias_list, municipio_keywords, freguesia_keywords, tipos_entidade_validos, limite_fuzzy):
    # Verificar se a coluna com o valor existe
    if coluna_verificar not in row:
        return None  # Sem a coluna a verificar, não há nada a validar

    # Obter o valor da coluna
    valor_verificado = str(row[coluna_verificar]) if pd.notna(row[coluna_verificar]) else None
    if not valor_verificado:
        return None  

    # Se a coluna entidade_type existir, usa-a; caso contrário, assume que pode ser um município ou freguesia
    entidade = normalize_text(row[entidade_type]) if entidade_type in row and pd.notna(row[entidade_type]) else None

    # Tentar validar como concelho
    if not entidade or entidade in municipio_keywords[:n_entidades_municipios]:
        resultado_concelho = validar_local(valor_verificado, concelhos_list, limite_fuzzy)
        if resultado_concelho:
            return resultado_concelho

    # Tentar validar como freguesia
    if not entidade or entidade in freguesia_keywords[:n_entidades_freguesias]:
        resultado_freguesia = validar_local(valor_verificado, freguesias_list, limite_fuzzy)
        if resultado_freguesia:
            return resultado_freguesia

    # Verificação genérica para município e freguesia
    resultado_generico = validar_generico_municipio_freguesia(valor_verificado)
    if resultado_generico:
        return resultado_generico

    # Se o tipo de entidade for válido, retorna o valor normalizado
    if entidade in tipos_entidade_validos:
        return clean_text(valor_verificado)

    # Se nada foi identificado, exibe um aviso e retorna None
    entidades_invalidas.append(row[coluna_verificar])
    return row[coluna_verificar]


# Aplicar a validação à coluna
df[coluna_verificar] = df.apply(
    lambda row: processar_entidade(
        row, coluna_verificar, entidade_type,
        concelhos_list, freguesias_list,
        municipio_keywords, freguesia_keywords,
        tipos_entidade_validos, limite_fuzzy
    ),
    axis=1
)

# Identificar linhas onde o resultado da validação falhou (valor final é inválido)
mask_invalidas = df[coluna_verificar].isin(entidades_invalidas)

if mask_invalidas.any():
    # Criar DataFrame com as linhas removidas e manter o valor original da entidade
    removidos = df[mask_invalidas].copy()
    removidos["MOTIVO REMOCAO"] = "ENTIDADE INVÁLIDA"

    # Reorganizar colunas com 'motivo_remocao' no início
    cols = ["MOTIVO REMOCAO"] + [col for col in removidos.columns if col != "MOTIVO REMOCAO"]
    removidos = removidos[cols]

    # Adicionar ao df_removidos
    df_removidos = pd.concat([df_removidos, removidos], ignore_index=True)

# Remover as linhas inválidas do DataFrame principal
df = df[~mask_invalidas]


#========================================================================================================#

#========================================== VALIDA RESPONSAVEL ==========================================#

if responsible in df.columns:
    if not df[responsible].notna().all() or not all(normalize_text(cell) in ["SIM", "NAO", "", " "] for cell in df[responsible]):
        df.drop(columns=[responsible], inplace=True)
        cols_targets = [col for col in cols_targets if col != responsible]  

#========================================================================================================#

#========================================== REMOVER DUPLICADOS ==========================================#
if coluna_verificar in df.columns:
    idxs_duplicated = set()

    for idx_v, valor in df[coluna_verificar].items():
        if pd.isna(valor):  
            continue

        valor_norm = normalize_text(str(valor)).replace(" ", "") 

        for idx_vc, valor_comparado in df[coluna_verificar].items():
            if (
                idx_v == idx_vc 
                or idx_v in idxs_duplicated 
                or idx_vc in idxs_duplicated 
                or pd.isna(valor_comparado)
            ):
                continue  

            valor_comparado_norm = normalize_text(str(valor_comparado)).replace(" ", "")

            if valor_norm == valor_comparado_norm:
                # Verifica se o valor é nulo OU está nos inválidos
                def contar_invalidos(idx):
                    return df.loc[idx].apply(lambda x: pd.isnull(x) or x in valores_invalidos).sum()

                if contar_invalidos(idx_v) <= contar_invalidos(idx_vc):
                    idxs_duplicated.add(idx_vc)
                else:
                    idxs_duplicated.add(idx_v)


    # Criar DataFrame dos valores removidos antes de excluí-los, com a coluna 'motivo_remocao'
    removidos = df.loc[list(idxs_duplicated)].copy()
    removidos["MOTIVO REMOCAO"] = "DUPLICADO"

    # Reorganizar colunas para garantir que 'motivo_remocao' fique em primeiro lugar
    cols = ["MOTIVO REMOCAO"] + [col for col in removidos.columns if col != "MOTIVO REMOCAO"]
    removidos = removidos[cols]

    # Adicionar as linhas removidas ao df_removidos
    df_removidos = pd.concat([df_removidos, removidos], ignore_index=True)

    # Remover duplicados e redefinir os índices
    df = df.drop(list(idxs_duplicated)).reset_index(drop=True)

else:
    print(f"A coluna '{coluna_verificar}' não existe no DataFrame.")


#========================================================================================================#

#========================================== COLUNA DE SUBMISSÃO =========================================#

# Identificar as colunas corretas usando match flexível
col_submissao = find_best_match(normalize_text(data_sub_key), df.columns)
col_ultima_acao = find_best_match(normalize_text(data_fim_key), df.columns)
col_inicio = find_best_match(normalize_text(data_inicio_key), df.columns)

df[col_submissao] = df[col_submissao].replace(valores_invalidos, pd.NA)

if col_ultima_acao:
    col_index = df.columns.get_loc(col_ultima_acao) + 1
else:
    col_index = df.columns.get_loc(col_submissao)

# Verificar se ambas as colunas foram encontradas
if col_submissao:

    # Criar a nova coluna com 'SIM' ou 'NÃO'    
    df.insert(col_index, col_sub, df[col_submissao].notna().map({True: "SIM", False: "NAO"}))

    # Inserir na posição correta em cols_targets
    if col_sub not in cols_targets:
        cols_targets.insert(col_index, col_sub)
else:
    print(f"Erro COLUNA DE SUBMISSÃO: A coluna '{data_sub_key}' não foi encontrada.")

#========================================================================================================#

#======================================== COLUNA TEMPO DE RESPOSTA ======================================#

# Verificar se as colunas foram encontradas
if col_inicio and col_ultima_acao:
    # Converter as colunas para datetime
    df[col_inicio] = pd.to_datetime(df[col_inicio], errors="coerce")
    df[col_ultima_acao] = pd.to_datetime(df[col_ultima_acao], errors="coerce")

    # Calcular a diferença entre as datas (em segundos)
    df["diferença segundos"] = (df[col_ultima_acao] - df[col_inicio]).dt.total_seconds()

    # Converter para o formato adequado (h:mm:ss ou mm:ss)
    df[col_temp] = df["diferença segundos"].apply(
        lambda x: f"{int(x // 3600):02}:{int((x % 3600) // 60):02}:{int(x % 60):02}" if pd.notna(x) and x >= 3600 else 
                  f"{int(x // 60):02}:{int(x % 60):02}" if pd.notna(x) else "00:00"
    )

    # Eliminar as linhas onde o TEMPO DE REALIZAÇÃO é "00:00"
    df = df[df[col_temp] != "00:00"]

    index = cols_targets.index(col_ultima_acao) + 1

    # Inserir a nova coluna col_temp na posição correta em cols_targets
    cols_targets.insert(index, col_temp)

else:
    print("Erro COLUNA TEMPO DE RESPOSTA: As colunas 'DATA DE INICIO' ou 'DATA DA ULTIMA ACCAO' não foram encontradas.")

#========================================================================================================#

#=========================================== COLUNAS DE CURSOS ==========================================#

# Dicionário para mapear as colunas
mapeamento_colunas = {}

# Identificar todas as colunas que contêm "Formação" ou "Curso" no nome
cols_formacao_curso = [col for col in df.columns if col and (any(normalize_text(palavra) in normalize_text(col) for palavra in palavras_chave_formacao))]

for col in cols_formacao_curso:
    col_normalizado = normalize_text(col)
    
    # Extrair o nome dentro de [ ]
    nome_encontrado = re.findall(r'\[(.*?)\]', col_normalizado)
    nome_final = " - ".join(nome_encontrado) if nome_encontrado else col_normalizado

    # Verificar se a coluna contém "comentário" ou "sugestões"
    if any(normalize_text(palavra) in col_normalizado for palavra in palavras_chave_comentario):
        nome_final = re.sub(r'(?i) - comentário', '', nome_final).strip()
        nome_final = f"{descricao_comentario}: {nome_final}"
        cols_comentarios_fc.append(col)

    # Verificar se a coluna contém "tempo do grupo"
    elif any(normalize_text(palavra) in col_normalizado for palavra in palavras_chave_tempo_grupo):
        nome_final = re.sub(r'(?i)\btempo do grupo:\s*', '', nome_final).strip()
        nome_final = f"{descricao_tempo_grupo}: {nome_final}"

    # Verificar se a coluna contém "áreas temáticas" no nome
    elif any(normalize_text(palavra) in col_normalizado for palavra in palavras_chave_areas_tematicas):
        nome_final = f"{descricao_interesse}: {nome_final}"
        cols_interesses_fc.append(col)

    # Verificar se a coluna contém "contínua"
    elif any(normalize_text(palavra) in col_normalizado for palavra in palavras_chave_continua):
        nome_final = f"{descricao_continua}: {nome_final}"
        cols_continua_fc.append(col)

    # Verificar se a coluna contém "preferência"
    elif any(normalize_text(palavra) in col_normalizado for palavra in palavras_chave_preferencia):
        nome_final = f"{descricao_preferencia}: {nome_final}"

    # Verificar se a coluna contém "regime"
    elif any(normalize_text(palavra) in col_normalizado for palavra in palavras_chave_regime):
        nome_final = f"{descricao_regime}: {nome_final}"

    else:
        nome_final = f"{descricao_formacao_curso}: {nome_final}"
        cols_sumformados.append(col)

    # Guardar no dicionário de mapeamento
    mapeamento_colunas[col] = nome_final

# Adicionar os nomes originais das colunas a cols_targets (mantendo os nomes no df)
if len(mapeamento_colunas) > 0:
    cols_targets += list(mapeamento_colunas.keys())

#========================================================================================================#

#====================================== PREENCHER CELULAS INVÁLIDAS =====================================#

def preencher_vazios(df):

    palavras_chave_comentario = [
    normalize_text(palavra) for palavra in [
        "comentário", "sugestões", "Temas não versados anteriormente"
    ]
]
    # Normalizar nomes de colunas
    colunas_o_que_preten_dem = [
        col for col in df.columns
        if any(palavra in normalize_text(col) for palavra in palavras_chave_comentario)
    ]


    # Verificar se há colunas correspondentes e preencher células vazias com "VAZIO"
    if colunas_o_que_preten_dem:
        for col in colunas_o_que_preten_dem:
            df[col] = df[col].apply(lambda x: "VAZIO" if pd.isna(x) or x in valores_invalidos else x)


    interesses_keys = [normalize_text(palavra) for palavra in palavras_chave_regime] + \
                  [normalize_text("INTERESSE"), normalize_text("FORMACAO CONTINUA"), normalize_text("RECETIVO A ACOLHER FORMANDOS")]

    colunas_interesse = [
        col for col in df.columns
        if any(palavra in col.upper() for palavra in  interesses_keys)
        and col not in colunas_o_que_preten_dem
    ]

    # Verificar se há colunas correspondentes e preencher células vazias com "NAO"
    if colunas_interesse:
        for col in colunas_interesse:
            df[col] = df[col].apply(lambda x: "NAO" if pd.isna(x) or x in valores_invalidos else x)
    
    return df


# Preencher colunas numéricas com 0
df[df.select_dtypes(include='number').columns] = df.select_dtypes(include='number').fillna(0)

# Preencher colunas cujo nome contém 'FORMACAO/CURSO' com 0
colunas_formacao_curso = [col for col in df.columns if any(palavra in col for palavra in ["NUMERO DE FORMANDOS", "TEMPO DO GRUPO"])]

# Função para tratar o valor "1 OU 2" e substituí-lo pelo valor mais elevado
def tratar_valor(x):
    if isinstance(x, str) and "OU" in x:
        # Extrair os números antes e depois de "OU", convertê-los em float e retornar o maior valor
        valores = x.split(" OU ")
        return max([float(val) for val in valores])  # Retorna o maior valor
    elif pd.isna(x) or x in valores_invalidos:
        return 0  # Substitui NaN ou valores inválidos por 0
    else:
        return x  # Mantém o valor original

# Aplicar a função nas colunas de interesse
for col in colunas_formacao_curso:
    df[col] = df[col].map(tratar_valor)



df = preencher_vazios(df)


#========================================================================================================#

#========================================== COLUNA Nº FORMANDOS =========================================#

# Criar uma nova coluna com a soma apenas dos valores inteiros
nformandos = df[cols_sumformados].apply(
    lambda row: sum(x for x in pd.to_numeric(row, errors='coerce').fillna(0) if x.is_integer()), 
    axis=1
)

# Adicionar a nova coluna ao DataFrame usando pd.concat
df = pd.concat([df, nformandos.rename(col_nformandos)], axis=1)

# Encontrar a última coluna numérica em cols_sumformados
ultima_coluna_numerica = None
for coluna in cols_sumformados:
    if pd.api.types.is_numeric_dtype(df[coluna]):  
        ultima_coluna_numerica = coluna

# Inserir o nome da nova coluna após a última coluna numérica
if ultima_coluna_numerica:
    index = cols_targets.index(ultima_coluna_numerica) + 1  
    cols_targets.insert(index, col_nformandos)


#========================================================================================================#

#=========================================== GUARDAR DATAFRAME ==========================================#

# Verificar quais colunas existem no DataFrame
cols_existentes = [col for col in cols_targets if col in df.columns]
cols_faltantes = set(cols_targets) - set(cols_existentes)

# Avisar sobre colunas em falta
if cols_faltantes:
    print(f"Aviso: As seguintes colunas não estão no DataFrame e serão ignoradas: {cols_faltantes}")

# Selecionar apenas as colunas que existem
if cols_existentes:
    df = df[cols_existentes]
else:
    print("Erro: Nenhuma das colunas alvo foi encontrada no DataFrame. O ETL pode falhar.")

# Aplicar o rename para substituir os nomes no DataFrame sem quebrar a estrutura
df.rename(columns=mapeamento_colunas, inplace=True)

# Guardar
df.to_excel(file_path_out, index=False)

# Abrir o ficheiro com openpyxl para modificar
wb = load_workbook(file_path_out)  
ws = wb.active  

# Alterar o título da aba ativa
ws.title = ws_title

# Ajustar o tamanho das colunas com base no tamanho dos cabeçalhos
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter 

    # Calcular a maior largura da coluna em uma única iteração
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    # Ajustar a largura da coluna
    ws.column_dimensions[col_letter].width = max_length + 2  

# Ajustar a altura da linha de cabeçalho (primeira linha)
ws.row_dimensions[1].height = 70  

# Centralizar o conteúdo de todas as células
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")  

# Salvar o ficheiro após as modificações
wb.save(file_path_out)

# Guardar as linhas removidas num ficheiro Excel separado
if not df_removidos.empty:
    df_removidos.to_excel(file_path_removed, index=False)

# Abrir o ficheiro salvo
if os.name == 'nt':  # Para sistemas Windows
    os.startfile(file_path_out)
elif os.name == 'posix':  # Para sistemas Unix (Linux/MacOS)
    if 'darwin' in subprocess.os.uname().sysname.lower():  # MacOS
        subprocess.call(['open', file_path_out])
    else:  # Linux
        subprocess.call(['xdg-open', file_path_out])

#========================================================================================================#

#=========================================== GUARDAR DATAFRAME ==========================================#

# Lista de palavras-chave
colunas_importantes = [
    "ENTIDADE DO SUBSETOR DA ADMINISTRAÇÃO LOCAL",
    "DESIGNAÇÃO DA ENTIDADE",
    "NUT II",
    "RESPONSÁVEL",
    "DATA DE INICIO",
    "DATA DA ULTIMA ACCAO",
    "O QUE PRETENDEM SOBRE",
    "TEMPO DO GRUPO",
    "INTERESSE",
    "FORMAÇÃO CONTÍNUA",
    "PREFERÊNCIA",
    "FORMAÇÃO/CURSO",
    "REGIME",
    "FOI SUBMETIDO?",
    "TEMPO DE REALIZAÇÃO",
    "Nº TOTAL DE FORMANDOS"
]

# Caminho do arquivo Excel
file_path_col = "C:/Users/franc/Documents/Estágio/codigo/ETL/teste_13_03/colunas_recolhidas/inqueritos_cols.xlsx"

# Função para contar as ocorrências das palavras-chave no DataFrame
def contar_ocorrencias(df):
    contagem_colunas = {}
    for palavra in colunas_importantes:
        contagem_colunas[palavra] = sum(df.columns.str.contains(palavra))
    return contagem_colunas

# Função para atualizar ou criar o arquivo Excel com os dados
def atualizar_excel(df, ano):
    # Contar as ocorrências das palavras-chave no df
    contagem_colunas = contar_ocorrencias(df)

    # Converter a contagem em um DataFrame
    contagem_df = pd.DataFrame([contagem_colunas])

    # Adicionar a coluna do ano (garantindo que o tipo seja numérico)
    contagem_df["ANO"] = int(ano)  # Garantir que o ano seja um número inteiro

    # Reorganizar a coluna "Ano" para ser a primeira
    contagem_df = contagem_df[["ANO"] + [col for col in contagem_df.columns if col != "ANO"]]

    # Se o arquivo já existir
    if os.path.exists(file_path_col):
        # Carregar o arquivo existente
        excel_df = pd.read_excel(file_path_col, sheet_name="Contagem")
        
        # Remover e substituir os dados do ano
        if int(ano) in excel_df["ANO"].values:
            excel_df = excel_df[excel_df["ANO"] != int(ano)]
            excel_df = pd.concat([excel_df, contagem_df], ignore_index=True)
        else:
            excel_df = pd.concat([excel_df, contagem_df], ignore_index=True)

    else:
        # Se o arquivo não existir, cria-lo com as colunas de contagem
        excel_df = contagem_df

    # Salvar o arquivo Excel
    with pd.ExcelWriter(file_path_col, engine='xlsxwriter') as writer:
        excel_df.to_excel(writer, sheet_name="Contagem", index=False)

# Função para extrair o ano do nome do ficheiro
def extrair_ano(nome_ficheiro):
    match = re.search(r"(20\d{2})", nome_ficheiro)
    return match.group(1) if match else "ANO_DESCONHECIDO"


ano_atual = extrair_ano(FILE_PATH_IN)


# Chamar a função para atualizar ou criar o arquivo Excel
atualizar_excel(df, ano_atual)

# Exibir a mensagem de sucesso
print(f"Os dados para o ano {ano_atual} foram atualizados ou adicionados com sucesso.")



#================================================= FIM ==================================================#
